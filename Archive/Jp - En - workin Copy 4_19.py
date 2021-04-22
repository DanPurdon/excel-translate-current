#!/usr/bin/env python3
#!print('Hello World!')

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment, Protection, Font
from copy import copy

inputFile = openpyxl.load_workbook('input_Japanese.xlsx')
rosettaStone = openpyxl.load_workbook('translation_Pairs_Unicode.xlsx')
outputFile = Workbook()

allInputSheetNames = inputFile.sheetnames
#allOutputSheetNames = outputFile.sheetnames

japanese = []
english = []

rosettaSheet = rosettaStone.active
outputSheet = outputFile.active
outputSheet.title = str(inputFile.sheetnames[0])


x = 1
while len(inputFile.sheetnames) > len(outputFile.sheetnames):
    outputFile.create_sheet(str(inputFile.sheetnames[x]))
    x += 1


def rosettaData ():
    i = 1
    i2 = 1
    while rosettaSheet.max_row >= i:
        japanese.append(rosettaSheet['A'+ str(i)].value)
        i += 1

    while rosettaSheet.max_row >= i2:
        english.append(rosettaSheet['B'+ str(i2)].value)
        i2 += 1

rosettaData()

def stylin (cellToCopy,cellToPaste):
    cellToPaste.font = copy(cellToCopy.font)
    cellToPaste.border = copy(cellToCopy.border)
    cellToPaste.fill = copy(cellToCopy.fill)
    cellToPaste.alignment = copy(cellToCopy.alignment)
    cellToPaste.number_format = copy(cellToCopy.number_format)

def scan ():
    for sheet in allInputSheetNames:
        currentSheet = inputFile[sheet]
        outputSheet = outputFile[sheet]
        for sheets in inputFile:
            for row in range(1, currentSheet.max_row + 1):
                for column in range(1, currentSheet.max_column + 1):


                    currentCell = currentSheet.cell(row=row, column=column)
                    outputCell = outputSheet.cell(row=row, column=column)
                    i = 0
                    while i < len(japanese):
                        if currentCell.value == japanese[i]:
                            outputSheet.cell(row=row, column=column, value=english[i])
                            stylin(currentCell, outputCell)
                            break

                        else:
                            i += 1

                    if outputCell.value is None:
                        outputSheet.cell(row=row, column=column, value=currentCell.value)
                        stylin(currentCell, outputCell)



scan()

outputFile.save('output_English.xlsx')
