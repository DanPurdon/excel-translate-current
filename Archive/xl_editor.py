#!/usr/bin/env python3
#!print('Hello World!')

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

inputFile = openpyxl.load_workbook('test_Translation.xlsx')
rosettaStone = openpyxl.load_workbook('test_Pairs_Unicode.xlsx')
outputFile = Workbook()

allInputSheetNames = inputFile.sheetnames
#allOutputSheetNames = outputFile.sheetnames

japanese = []
english = []

rosettaSheet = rosettaStone.active
outputSheet = outputFile.active
outputSheet.title = str(inputFile.sheetnames[0])

while len(inputFile.sheetnames) > len(outputFile.sheetnames):
    x = 1
    outputFile.create_sheet(str(inputFile.sheetnames[x]))
    x += 1




def rosettaData ():
    i = 1
    i2 = 1
    while rosettaSheet.max_row >= i:
        japanese.append(rosettaSheet['A'+ str(i)].value)
        i += 1

    while rosettaSheet.max_row >= i2:
        english.append(rosettaSheet['B'+ str(i2)].value)
        i2 += 1

rosettaData()



def scan ():
    for sheet in allInputSheetNames:
        currentSheet = inputFile[sheet]
        outputSheet = outputFile[sheet]
        for sheets in inputFile:
            for row in range(1, currentSheet.max_row + 1):
                for column in range(1, currentSheet.max_column + 1):


                    currentCell = currentSheet.cell(row=row, column=column)
                    outputCell = outputSheet.cell(row=row, column=column)
                    i = 0
                    while i < len(japanese):
                        if currentCell.value == japanese[i]:
                            outputSheet.cell(row=row, column=column, value=english[i])
                            break
                        else:
                            i += 1

                    if outputCell.value is None:
                        outputSheet.cell(row=row, column=column, value=currentCell.value)


scan()

outputFile.save('test_Output.xlsx')

#print("All sheet names {} " .format(inputFile.sheetnames))

#for sheet in allInputSheetNames:
    #print("Current sheet name is {}" .format(sheet))
    #currentSheet = inputFile[sheet]
    # print(currentSheet['B4'].value)

    #print max numbers of wors and colums for each sheet
    #print(currentSheet.max_row)
    #print(currentSheet.max_column)

    #for row in range(1, currentSheet.max_row + 1):
        #print(row)
        #for column in "ABCDEF":  # Here you can add or reduce the columns
            #cell_name = "{}{}".format(column, row)
            #print(cell_name)
            #print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))
