
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color
from copy import copy
import sys
recursion = 15000
sys.setrecursionlimit(recursion)

inputJapanese = openpyxl.load_workbook('input_Farm_Japanese.xlsx')
inputEnglish = openpyxl.load_workbook('input_Farm_English.xlsx')
outputFile = Workbook()

allJapaneseInputSheetNames = inputJapanese.sheetnames
allEnglishInputSheetNames = inputEnglish.sheetnames
outputSheet = outputFile.active



def scan_Japanese ():
    i = 1
    for sheet in allJapaneseInputSheetNames:
        currentJapaneseSheet = inputJapanese[sheet]

        for sheets in inputJapanese:
            for row in range(1, currentJapaneseSheet.max_row + 1):
                for column in range(1, currentJapaneseSheet.max_column + 1):
                    currentJapaneseCell = currentJapaneseSheet.cell(row=row, column=column)

                    if currentJapaneseCell.value is None:
                        continue
                    elif currentJapaneseCell.value == "":
                        continue
                    else:
                        outputSheet.cell(row=i, column=1, value=currentJapaneseCell.value)
                        i += 1

scan_Japanese()

def scan_English ():
    i = 1
    for sheet in allEnglishInputSheetNames:
        currentEnglishSheet = inputEnglish[sheet]

        for sheets in inputEnglish:
            for row in range(1, currentEnglishSheet.max_row + 1):
                for column in range(1, currentEnglishSheet.max_column + 1):
                    currentEnglishCell = currentEnglishSheet.cell(row=row, column=column)

                    if currentEnglishCell.value is None:
                        continue
                    elif currentEnglishCell.value == "":
                        continue
                    else:
                        outputSheet.cell(row=i, column=2, value=currentEnglishCell.value)
                        i += 1

scan_English()



def scan_Dupes():
    check = outputSheet.max_row
    while check > 1:
        compare = check - 1
        while compare > 0:
            if outputSheet['A' + str(check)].value == outputSheet['A' + str(compare)].value:
                outputSheet.delete_rows(check)
                break
            compare -= 1
        check -= 1



scan_Dupes()

outputFile.save('output_Farm.xlsx')
