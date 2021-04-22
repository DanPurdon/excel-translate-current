#!/usr/bin/env python3
#!print('Hello World!')

import tkinter as tk
from tkinter import filedialog as fd
from tkinter import *
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment, Protection, Font
from copy import copy


rosettaStone = openpyxl.load_workbook('translation_Pairs_Unicode.xlsx')
outputFile = Workbook()
# inputFile = openpyxl.load_workbook('input_Japanese.xlsx')

# Rosetta translation pair prep
japanese = []
english = []

rosettaSheet = rosettaStone.active

def rosettaData ():
    i = 1
    i2 = 1
    while rosettaSheet.max_row >= i:
        japanese.append(rosettaSheet['A'+ str(i)].value)
        i += 1

    while rosettaSheet.max_row >= i2:
        english.append(rosettaSheet['B'+ str(i2)].value)
        i2 += 1
rosettaData()
# /Rosetta translation pair list

# Cell style copier
def stylin (cellToCopy,cellToPaste):
    cellToPaste.font = copy(cellToCopy.font)
    cellToPaste.border = copy(cellToCopy.border)
    cellToPaste.fill = copy(cellToCopy.fill)
    cellToPaste.alignment = copy(cellToCopy.alignment)
    cellToPaste.number_format = copy(cellToCopy.number_format)
# /cell styler


window = tk.Tk()
window.geometry("265x185")
window.title("DanCo Translate")
greeting = tk.Label(text="DanCo Translation Widget", borderwidth = 10, width = 30)
greeting.pack(pady=7)

# Radio Buttons
language = Radiobutton ()
langChoice = IntVar()
langChoice.set(1)
R1 = Radiobutton(window, text="Japanese -> English", variable=langChoice, value=1)
R1.pack()

R2 = Radiobutton(window, text="English -> Japanese", variable=langChoice, value=2)
R2.pack(padx=0, pady=0)
# /Radio Buttons


spacerFrame = tk.Frame(master=window, height=7)
spacerFrame.pack()

# Directory Buttons
stringVar1 = StringVar()
stringVar1.set("Select file to translate")
stringVar2 = StringVar()

def filePath():
    path = fd.askopenfilename(initialdir = "/",title = "Select file",filetypes = (("Excel files","*.xlsx;*.xls"),("all files","*.*")))
    if path == "":
        pass
    elif langChoice.get() == 1:
        scanJapanese(path)
        outputFile.save('output_English.xlsx')
        stringVar1.set("Saved output_English.xlsx to: ")
        stringVar2.set(str(path))
    else:
        scanEnglish(path)
        outputFile.save('output_Japanese.xlsx')
        stringVar1.set("Saved output_Japanese.xlsx to: ")
        stringVar2.set(str(path))

buttonFrame = tk.Frame(height = 15)
tk.Button(text='Select file',command=filePath,master=buttonFrame, borderwidth = 2).grid(column = 0, row = 0, padx=4, pady=2, sticky = "s")
tk.Button(text='Close',command=window.destroy,master=buttonFrame, borderwidth = 2).grid(column = 1, row = 0, padx=4, pady=2, sticky = "s")
buttonFrame.pack()
# /Directory buttons



guideText1 = tk.Label(textvariable = stringVar1)
guideText1.pack()
guideText2 = tk.Label(textvariable = stringVar2)
guideText2.pack()


def scanJapanese (path):
    inputFile = openpyxl.load_workbook(path)
    allInputSheetNames = inputFile.sheetnames
    outputSheet = outputFile.active
    outputSheet.title = str(inputFile.sheetnames[0])
    x = 1
    while len(inputFile.sheetnames) > len(outputFile.sheetnames):
        outputFile.create_sheet(str(inputFile.sheetnames[x]))
        x += 1
    for sheet in allInputSheetNames:
        currentSheet = inputFile[sheet]
        outputSheet = outputFile[sheet]
        for sheets in inputFile:
            for row in range(1, currentSheet.max_row + 1):
                for column in range(1, currentSheet.max_column + 1):


                    currentCell = currentSheet.cell(row=row, column=column)
                    outputCell = outputSheet.cell(row=row, column=column)
                    i = 0
                    while i < len(japanese):
                        if currentCell.value == japanese[i]:
                            outputSheet.cell(row=row, column=column, value=english[i])
                            stylin(currentCell, outputCell)
                            break

                        else:
                            i += 1

                    if outputCell.value is None:
                        outputSheet.cell(row=row, column=column, value=currentCell.value)
                        stylin(currentCell, outputCell)


def scanEnglish (path):
    inputFile = openpyxl.load_workbook(path)
    allInputSheetNames = inputFile.sheetnames
    outputSheet = outputFile.active
    outputSheet.title = str(inputFile.sheetnames[0])
    x = 1
    while len(inputFile.sheetnames) > len(outputFile.sheetnames):
        outputFile.create_sheet(str(inputFile.sheetnames[x]))
        x += 1
    for sheet in allInputSheetNames:
        currentSheet = inputFile[sheet]
        outputSheet = outputFile[sheet]
        for sheets in inputFile:
            for row in range(1, currentSheet.max_row + 1):
                for column in range(1, currentSheet.max_column + 1):


                    currentCell = currentSheet.cell(row=row, column=column)
                    outputCell = outputSheet.cell(row=row, column=column)
                    i = 0
                    while i < len(english):
                        if currentCell.value == english[i]:
                            outputSheet.cell(row=row, column=column, value=japanese[i])
                            stylin(currentCell, outputCell)
                            break

                        else:
                            i += 1

                    if outputCell.value is None:
                        outputSheet.cell(row=row, column=column, value=currentCell.value)
                        stylin(currentCell, outputCell)



window.mainloop()
